from __future__ import annotations

import io


class ExtractionError(Exception):
    """Raised when a supported-format file cannot be parsed."""


class UnsupportedFileTypeError(ValueError):
    """Raised when the MIME type is not in SUPPORTED_MIME_TYPES."""


class ExtractionService:
    MAX_FILE_SIZE: int = 10 * 1024 * 1024  # 10 MB

    SUPPORTED_MIME_TYPES: frozenset[str] = frozenset({
        "application/pdf",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "text/plain",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    })

    def extract(self, filename: str, content_type: str, data: bytes) -> str:
        """Dispatch to the appropriate parser based on content_type.

        Returns a plain-text string (may be empty).
        Raises UnsupportedFileTypeError for unknown MIME types.
        Raises ExtractionError for parse failures.
        """
        if content_type == "text/plain":
            return self._extract_txt(data)
        elif content_type == "application/pdf":
            return self._extract_pdf(data)
        elif content_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
            return self._extract_docx(data)
        elif content_type in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        ):
            return self._extract_excel(data)
        else:
            raise UnsupportedFileTypeError(
                f"Unsupported file type: {content_type}. "
                "Accepted types: pdf, docx, txt, xls, xlsx."
            )

    def _extract_txt(self, data: bytes) -> str:
        return data.decode("utf-8", errors="replace")

    def _extract_pdf(self, data: bytes) -> str:
        try:
            from pypdf import PdfReader

            reader = PdfReader(io.BytesIO(data))
            parts: list[str] = []
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    parts.append(text)
            return "\n".join(parts)
        except Exception as exc:
            raise ExtractionError(f"Failed to parse PDF: {exc}") from exc

    def _extract_docx(self, data: bytes) -> str:
        try:
            from docx import Document

            doc = Document(io.BytesIO(data))
            parts = [para.text for para in doc.paragraphs if para.text]
            return "\n".join(parts)
        except Exception as exc:
            raise ExtractionError(f"Failed to parse DOCX: {exc}") from exc

    def _extract_excel(self, data: bytes) -> str:
        """
        Extract Excel content preserving row/column structure.
        Each row becomes one line; columns are separated by a tab character.
        This allows the deduplication service to treat each line as a full row.
        """
        try:
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(data), read_only=True)
            lines: list[str] = []
            ws = wb.worksheets[0]   # process first sheet only
            for row in ws.iter_rows(values_only=True):
                # Convert each cell to string; None → empty string
                cells = [str(cell) if cell is not None else "" for cell in row]
                # Skip entirely empty rows
                if any(c.strip() for c in cells):
                    lines.append("\t".join(cells))
            return "\n".join(lines)
        except Exception as exc:
            raise ExtractionError(f"Failed to parse Excel file: {exc}") from exc


extraction_service = ExtractionService()
