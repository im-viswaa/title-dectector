"""
Excel-native deduplication service.

Accepts raw Excel bytes (.xlsx / .xls), removes duplicate rows while preserving:
- All columns and their original values (original text is never modified)
- Header row (row 1 is always kept)
- Original row order (first occurrence wins)

Duplicate detection strategy:
1. Find the "title" column — the first column whose header contains a keyword
   like "title", "name", "topic", "subject", or "content" (case-insensitive).
2. If no such column is found, use ALL columns for comparison.
3. Normalise the comparison value: strip leading numbering ("1. ", "6. "),
   collapse whitespace, lowercase.

This means "1. Introduction to Python" and "6. Introduction to Python"
are treated as duplicates even if they have different IDs.
"""
from __future__ import annotations

import io
import re
from typing import Any

# Matches leading list numbering: "1. " "12. " "(3) " "3) " "a) " "A. "
_LEADING_NUMBER_RE = re.compile(
    r"^\s*(?:\d+[\.\)]\s*|[a-zA-Z]{1,4}[\.\)]\s*|\([^)]{1,4}\)\s*)"
)

# Column header keywords that indicate a "title" column
_TITLE_KEYWORDS = ("title", "name", "topic", "subject", "content", "heading", "label")


def _normalize_value(value: Any) -> str:
    """Normalise a cell value for duplicate comparison."""
    if value is None:
        return ""
    text = str(value).strip()
    text = _LEADING_NUMBER_RE.sub("", text).strip()
    return re.sub(r"\s+", " ", text).lower()


def _find_title_column(headers: list[Any]) -> int | None:
    """Return the index of the first column whose header matches a title keyword."""
    for i, h in enumerate(headers):
        if h is not None and any(kw in str(h).lower() for kw in _TITLE_KEYWORDS):
            return i
    return None


def deduplicate_excel_bytes(data: bytes) -> tuple[bytes, int, int]:
    """
    Remove duplicate rows from an Excel workbook.

    Returns:
        (output_bytes, original_data_row_count, duplicates_removed)

    - Row 1 (header) is always kept.
    - Duplicates are detected using the title column if one exists,
      otherwise all columns are used.
    - Original cell values are never modified.
    """
    from openpyxl import load_workbook, Workbook

    wb_in = load_workbook(io.BytesIO(data))
    ws_in = wb_in.active

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = ws_in.title  # type: ignore[assignment]

    seen: set[str] = set()
    original_count = 0
    duplicates_removed = 0
    header_written = False
    title_col_idx: int | None = None

    for row in ws_in.iter_rows(values_only=True):
        row_values: list[Any] = list(row)

        # ── Header row ────────────────────────────────────────────────────────
        if not header_written:
            ws_out.append(row_values)
            header_written = True
            title_col_idx = _find_title_column(row_values)
            continue

        original_count += 1

        # Preserve fully empty rows without counting as duplicates
        if all(_normalize_value(v) == "" for v in row_values):
            ws_out.append(row_values)
            continue

        # Build comparison key
        if title_col_idx is not None and title_col_idx < len(row_values):
            # Use only the title column for duplicate detection
            key = _normalize_value(row_values[title_col_idx])
        else:
            # No title column found — use all columns
            key = "\x00".join(_normalize_value(v) for v in row_values)

        if key in seen:
            duplicates_removed += 1
        else:
            seen.add(key)
            ws_out.append(row_values)   # original values preserved

    buf = io.BytesIO()
    wb_out.save(buf)
    return buf.getvalue(), original_count, duplicates_removed
